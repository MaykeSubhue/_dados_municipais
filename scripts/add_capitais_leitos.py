from __future__ import annotations

import json
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = ROOT / "tmp" / "demanda_comparacao_municipios.xlsx"
PDF_PATH = ROOT / "tmp" / "capitais_leitos_por_habitante.pdf"
DOWNLOADS_PDF = (
    Path.home()
    / "Downloads"
    / "Demanda_de_dados_MRJ_Municipios_ERJ.xlsx - Capitais - Leitos por habitante.pdf"
)
TSV_PATH = ROOT / "tmp" / "capitais_leitos_por_1000.tsv"
SOURCE_FILE_LABEL = "Demanda_de_dados_MRJ_Municipios_ERJ.xlsx - Capitais - Leitos por habitante.pdf"

TARGET_SHEET = "Capitais - Leitos por 1.000 habitante"
TEMPLATE_SHEET = "Leitos por 1.000 hab."
ORDER_TEMPLATE = "Capitais - AIH"
YEARS = [2020, 2021, 2022, 2023, 2024, 2025]

UF_BY_CAPITAL = {
    "Aracaju": "SE",
    "Belo Horizonte": "MG",
    "Belém": "PA",
    "Boa Vista": "RR",
    "Brasília": "DF",
    "Campo Grande": "MS",
    "Cuiabá": "MT",
    "Curitiba": "PR",
    "Florianópolis": "SC",
    "Fortaleza": "CE",
    "Goiânia": "GO",
    "João Pessoa": "PB",
    "Macapá": "AP",
    "Maceió": "AL",
    "Manaus": "AM",
    "Natal": "RN",
    "Palmas": "TO",
    "Porto Alegre": "RS",
    "Porto Velho": "RO",
    "Recife": "PE",
    "Rio Branco": "AC",
    "Rio de Janeiro": "RJ",
    "Salvador": "BA",
    "São Luís": "MA",
    "São Paulo": "SP",
    "Teresina": "PI",
    "Vitória": "ES",
}


def source_pdf() -> Path:
    if PDF_PATH.exists():
        return PDF_PATH
    if DOWNLOADS_PDF.exists():
        return DOWNLOADS_PDF
    raise FileNotFoundError(f"PDF não encontrado: {PDF_PATH} ou {DOWNLOADS_PDF}")


def as_float(value: Any) -> float:
    if value is None:
        raise ValueError("Valor vazio")
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(".", "").replace(",", ".")
    return float(text)


def parse_pdf_rows(pdf_path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                for source_row in table[3:]:
                    capital = str(source_row[0] or "").strip()
                    if not capital or capital.startswith("TOTAL"):
                        continue
                    if "Capitais" in capital:
                        continue
                    if capital not in UF_BY_CAPITAL:
                        raise ValueError(f"Capital inesperada no PDF: {capital!r}")
                    values = [as_float(source_row[col]) for col in range(13, 19)]
                    rows.append(
                        {
                            "capital": capital,
                            "location": f"{capital} ({UF_BY_CAPITAL[capital]})",
                            "values": dict(zip(YEARS, values)),
                        }
                    )

    if len(rows) != 27:
        raise ValueError(f"Esperadas 27 capitais no PDF; encontradas {len(rows)}")
    return rows


def clone_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.number_format:
        target.number_format = source.number_format


def ordered_rows(wb, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_location = {row["location"]: row for row in rows}
    if ORDER_TEMPLATE not in wb.sheetnames:
        return rows
    ordered: list[dict[str, Any]] = []
    template = wb[ORDER_TEMPLATE]
    for row_index in range(4, template.max_row + 1):
        location = template.cell(row=row_index, column=1).value
        if location in by_location:
            ordered.append(by_location.pop(location))
    ordered.extend(sorted(by_location.values(), key=lambda item: item["capital"]))
    return ordered


def recreate_sheet(wb):
    if TARGET_SHEET in wb.sheetnames:
        del wb[TARGET_SHEET]
    template = wb[TEMPLATE_SHEET]
    target_index = wb.sheetnames.index(TEMPLATE_SHEET) + 1
    ws = wb.create_sheet(TARGET_SHEET, target_index)

    for row in range(1, 38):
        ws.row_dimensions[row].height = template.row_dimensions[row].height
        for col in range(1, 10):
            clone_cell_style(template.cell(row=row, column=min(col, template.max_column)), ws.cell(row=row, column=col))

    for col in range(1, 10):
        letter = get_column_letter(col)
        source_letter = get_column_letter(min(col, template.max_column))
        ws.column_dimensions[letter].width = template.column_dimensions[source_letter].width

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["I"].width = 42
    ws.freeze_panes = template.freeze_panes
    ws.sheet_view.showGridLines = template.sheet_view.showGridLines
    ws.merge_cells("A1:G1")
    ws.merge_cells("B2:G2")
    return ws


def write_sheet(ws, rows: list[dict[str, Any]], pdf_path: Path) -> None:
    ws["A1"] = "Leitos hospitalares por 1.000 habitantes — Capitais do Brasil, 2020–2025"
    ws["A2"] = "Capital"
    ws["B2"] = "Leitos por 1.000 hab."
    for col, year in enumerate(YEARS, start=2):
        ws.cell(row=3, column=col, value=year)

    for row_index, item in enumerate(rows, start=4):
        ws.cell(row=row_index, column=1, value=item["location"])
        for col, year in enumerate(YEARS, start=2):
            cell = ws.cell(row=row_index, column=col, value=item["values"][year])
            cell.number_format = "#,##0.00"

    notes = [
        "Fonte e notas",
        f"Fonte: {SOURCE_FILE_LABEL}.",
        "Indicador: leitos hospitalares por capital, normalizados por 1.000 habitantes.",
        "Período: 2020 a 2025. O PDF traz população, total de leitos e a razão final por 1.000 habitantes.",
        "Fórmula da fonte: leitos / população do ano x 1.000.",
        "Valores preenchidos a partir da seção 'Leitos por 1.000 hab.' do PDF.",
    ]
    for offset, note in enumerate(notes, start=6):
        ws.cell(row=offset, column=9, value=note)


def write_tsv(rows: list[dict[str, Any]], pdf_path: Path) -> None:
    table: list[list[Any]] = [
        ["Leitos hospitalares por 1.000 habitantes — Capitais do Brasil, 2020–2025", "", "", "", "", "", "", "", ""],
        ["Capital", "Leitos por 1.000 hab.", "", "", "", "", "", "", ""],
        ["", *YEARS, "", ""],
    ]
    for item in rows:
        table.append([item["location"], *[item["values"][year] for year in YEARS], "", ""])
    while len(table) < 6:
        table.append([""] * 9)
    notes = [
        "Fonte e notas",
        f"Fonte: {SOURCE_FILE_LABEL}.",
        "Indicador: leitos hospitalares por capital, normalizados por 1.000 habitantes.",
        "Período: 2020 a 2025. O PDF traz população, total de leitos e a razão final por 1.000 habitantes.",
        "Fórmula da fonte: leitos / população do ano x 1.000.",
        "Valores preenchidos a partir da seção 'Leitos por 1.000 hab.' do PDF.",
    ]
    for index, note in enumerate(notes, start=5):
        while len(table) <= index:
            table.append([""] * 9)
        table[index][8] = note

    def cell_text(value: Any) -> str:
        if isinstance(value, float):
            return f"{value:.12g}".replace(".", ",")
        return "" if value is None else str(value)

    TSV_PATH.write_text(
        "\n".join("\t".join(cell_text(value) for value in row) for row in table),
        encoding="utf-8",
    )


def main() -> None:
    pdf_path = source_pdf()
    rows = parse_pdf_rows(pdf_path)
    wb = load_workbook(WORKBOOK_PATH)
    rows = ordered_rows(wb, rows)
    ws = recreate_sheet(wb)
    write_sheet(ws, rows, pdf_path)
    wb.save(WORKBOOK_PATH)
    write_tsv(rows, pdf_path)

    rio = next(row for row in rows if row["location"] == "Rio de Janeiro (RJ)")
    summary = {
        "updatedAt": datetime.now().isoformat(timespec="seconds"),
        "source": str(pdf_path),
        "workbook": str(WORKBOOK_PATH),
        "sheet": TARGET_SHEET,
        "rows": len(rows),
        "tsv": str(TSV_PATH),
        "rioSample": rio["values"],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
