from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from statistics import mean
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_XLSX = ROOT.parent / "tmp" / "demanda_comparacao_municipios.xlsx"
OUTPUT_JSON = ROOT / "data.json"
DEFAULT_YEARS = [2020, 2021, 2022, 2023, 2024, 2025]
SCCS_SHEETS = {"Produção por 100mil habitantes", "OCI por 100 mil habitantes"}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def slugify(value: str) -> str:
    replacements = {
        "á": "a",
        "à": "a",
        "ã": "a",
        "â": "a",
        "é": "e",
        "ê": "e",
        "í": "i",
        "ó": "o",
        "ô": "o",
        "õ": "o",
        "ú": "u",
        "ü": "u",
        "ç": "c",
    }
    text = value.lower()
    for source, target in replacements.items():
        text = text.replace(source, target)
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text or "indicador"


def as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("R$", "").replace("\xa0", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def infer_scope(sheet_name: str) -> str:
    if sheet_name in SCCS_SHEETS:
        return "SCCS"
    return "Capitais do Brasil" if sheet_name.startswith("Capitais -") else "Municípios do ERJ"


def infer_unit(title: str, subtitle: str) -> str:
    title_l = title.lower()
    subtitle_l = subtitle.lower()
    if "r$" in title_l or "r$" in subtitle_l or "valor" in title_l:
        return "R$ por 1.000 hab."
    if "100.000" in title or "100.000" in subtitle:
        return "por 100.000 nascidos vivos"
    if "tempo médio" in title_l:
        return "dias"
    if "taxa de natalidade" in title_l:
        return "por 1.000 hab."
    if "100 mil" in title_l or "100 mil" in subtitle_l:
        return "por 100 mil hab."
    if "1.000" in title or "1.000" in subtitle or "1.000" in subtitle_l:
        return "por 1.000 hab."
    return subtitle or "valor"


def year_columns(ws: Any) -> list[tuple[int, int]]:
    columns: list[tuple[int, int]] = []
    for col_index in range(2, ws.max_column + 1):
        raw_value = ws.cell(3, col_index).value
        value = clean_text(raw_value)
        if isinstance(raw_value, (int, float)) and 1900 <= int(raw_value) <= 2100:
            columns.append((col_index, int(raw_value)))
        elif value.isdigit() and 1900 <= int(value) <= 2100:
            columns.append((col_index, int(value)))
    return columns


def notes_from_sheet(ws: Any, first_data_row: int) -> list[str]:
    notes: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        for idx in [0, 8, 9]:
            if idx < len(row):
                text = clean_text(row[idx])
                if text and text.lower() not in {"fonte e notas", "município", "capital"}:
                    if idx == 0 and row[1] is not None:
                        continue
                    if text not in notes and (
                        "fonte" in text.lower()
                        or "dados" in text.lower()
                        or "população" in text.lower()
                        or "razão" in text.lower()
                        or "período" in text.lower()
                        or "situação" in text.lower()
                        or "fórmula" in text.lower()
                        or "total" in text.lower()
                        or idx in {8, 9}
                    ):
                        notes.append(text)
    return notes[:10]


def parse_sheet(ws: Any) -> dict[str, Any] | None:
    title = clean_text(ws.cell(1, 1).value)
    label = clean_text(ws.cell(2, 2).value)
    if not title or not label:
        return None
    columns = year_columns(ws)
    if not columns:
        columns = [(col_offset, year) for col_offset, year in enumerate(DEFAULT_YEARS, start=2)]

    rows: list[dict[str, Any]] = []
    records: list[dict[str, Any]] = []
    for row_idx in range(4, ws.max_row + 1):
        location = clean_text(ws.cell(row_idx, 1).value)
        if not location:
            continue
        lower = location.lower()
        if lower.startswith("fonte") or lower.startswith("total"):
            continue
        if location in {"Município", "Capital"}:
            continue

        values: dict[str, float] = {}
        has_year = False
        for col_index, year in columns:
            value = as_float(ws.cell(row_idx, col_index).value)
            if value is None:
                continue
            values[str(year)] = value
            has_year = True
            records.append(
                {
                    "scope": infer_scope(ws.title),
                    "indicator": ws.title,
                    "location": location,
                    "year": year,
                    "value": value,
                }
            )
        if has_year:
            rows.append({"location": location, "values": values})

    if not rows:
        return None

    indicator_years = sorted({int(year) for row in rows for year in row["values"]})
    all_values = [value for row in rows for value in row["values"].values()]
    zero_count = sum(1 for value in all_values if value == 0)
    non_zero_count = len(all_values) - zero_count
    zero_share = zero_count / len(all_values) if all_values else 0
    values_2025 = [row["values"].get("2025") for row in rows if "2025" in row["values"]]
    non_zero_2025 = [value for value in values_2025 if value and value > 0]

    return {
        "id": slugify(ws.title),
        "sheet": ws.title,
        "title": title,
        "label": label,
        "scope": infer_scope(ws.title),
        "unit": infer_unit(title, label),
        "years": indicator_years,
        "rare": zero_share >= 0.45,
        "rows": rows,
        "notes": notes_from_sheet(ws, 4),
        "stats": {
            "locations": len(rows),
            "records": len(all_values),
            "zeroRecords": zero_count,
            "nonZeroRecords": non_zero_count,
            "zeroShare": zero_share,
            "latestNonZero": len(non_zero_2025),
            "latestAverage": mean(non_zero_2025) if non_zero_2025 else 0,
        },
        "records": records,
    }


def main() -> None:
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    indicators = [parsed for ws in wb.worksheets if (parsed := parse_sheet(ws))]
    records = [record for indicator in indicators for record in indicator["records"]]
    for indicator in indicators:
        indicator.pop("records", None)

    years = sorted({year for indicator in indicators for year in indicator["years"]}) or DEFAULT_YEARS
    payload = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "source": {
            "title": "Demanda_Comparação Municipios",
            "spreadsheetUrl": "https://docs.google.com/spreadsheets/d/18MPMdnP5OQ4q39GwBh5BIxLaqakhrtUHDhFPRkWDCNw/edit",
        },
        "years": years,
        "indicators": indicators,
        "records": records,
        "summary": {
            "indicatorCount": len(indicators),
            "recordCount": len(records),
            "scopeCounts": {
                scope: sum(1 for item in indicators if item["scope"] == scope)
                for scope in sorted({item["scope"] for item in indicators})
            },
            "locationCounts": {
                scope: len(
                    {
                        row["location"]
                        for item in indicators
                        if item["scope"] == scope
                        for row in item["rows"]
                    }
                )
                for scope in sorted({item["scope"] for item in indicators})
            },
        },
    }
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT_JSON} with {len(indicators)} indicators and {len(records)} records")


if __name__ == "__main__":
    main()
