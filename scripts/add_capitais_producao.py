from __future__ import annotations

import json
import re
import urllib.request
from copy import copy
from datetime import datetime
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = ROOT / "tmp" / "demanda_comparacao_municipios.xlsx"
PANEL_DATA_PATH = ROOT / "painel-saude" / "data.json"
YEARS = [2020, 2021, 2022, 2023, 2024, 2025]

INPUTS = [
    {
        "sheet": "Capitais - Produção Cirurgica",
        "pdf": Path(r"C:\Users\Usuário\Downloads\Demanda_de_dados_MRJ_Municipios_ERJ.xlsx - Producao Cirurgica - Capitais.pdf"),
        "title": "Produção cirúrgica por 1.000 habitantes - Capitais do Brasil, 2020-2025",
        "label": "Procedimentos cirúrgicos por 1.000 hab.",
        "source": "Qtd.aprovada por Capital e Ano atendimento / Grupo procedimento: 04 Procedimentos cirurgicos",
    },
    {
        "sheet": "Capitais - Produção Clinica",
        "pdf": Path(r"C:\Users\Usuário\Downloads\Demanda_de_dados_MRJ_Municipios_ERJ.xlsx - Producao Clinica - Capitais.pdf"),
        "title": "Produção clínica por 1.000 habitantes - Capitais do Brasil, 2020-2025",
        "label": "Procedimentos clínicos por 1.000 hab.",
        "source": "Qtd.aprovada por Capital e Ano atendimento / Grupo procedimento: 03 Procedimentos clinicos",
    },
]


UF_BY_PREFIX = {
    "11": "RO",
    "12": "AC",
    "13": "AM",
    "14": "RR",
    "15": "PA",
    "16": "AP",
    "17": "TO",
    "21": "MA",
    "22": "PI",
    "23": "CE",
    "24": "RN",
    "25": "PB",
    "26": "PE",
    "27": "AL",
    "28": "SE",
    "29": "BA",
    "31": "MG",
    "32": "ES",
    "33": "RJ",
    "35": "SP",
    "41": "PR",
    "42": "SC",
    "43": "RS",
    "50": "MS",
    "51": "MT",
    "52": "GO",
    "53": "DF",
}


def sidra_json(url: str) -> list[dict[str, str]]:
    with urllib.request.urlopen(url, timeout=60) as response:
        return json.load(response)


def parse_number(value: str) -> int:
    return int(value.replace(".", "").replace(",", ""))


def parse_pdf_rows(pdf_path: Path) -> list[dict[str, object]]:
    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    rows: list[dict[str, object]] = []
    pattern = re.compile(r"^(\d{6})\s+(.+?)\s+((?:[\d.]+\s+){6}[\d.]+)$")
    for line in text.splitlines():
        match = pattern.match(line.strip())
        if not match:
            continue
        code6, name, values = match.groups()
        numbers = values.split()
        if len(numbers) != 7:
            raise ValueError(f"Unexpected year count in line: {line}")
        yearly = {year: parse_number(numbers[index]) for index, year in enumerate(YEARS)}
        uf = UF_BY_PREFIX[code6[:2]]
        rows.append(
            {
                "code6": code6,
                "name": name,
                "location": f"{name} ({uf})",
                "raw": yearly,
            }
        )

    if len(rows) != 27:
        raise ValueError(f"Expected 27 capitals in {pdf_path.name}, found {len(rows)}")
    return rows


def fetch_code_map(code6_values: set[str]) -> dict[str, str]:
    data = sidra_json("https://apisidra.ibge.gov.br/values/t/6579/n6/all/v/9324/p/2025?formato=json")
    code_map: dict[str, str] = {}
    for row in data[1:]:
        code7 = str(row["D1C"])
        prefix = code7[:6]
        if prefix in code6_values:
            code_map[prefix] = code7
    missing = sorted(code6_values - set(code_map))
    if missing:
        raise ValueError(f"Missing SIDRA municipality codes for DATASUS prefixes: {missing}")
    return code_map


def fetch_population(code7_values: list[str]) -> dict[str, dict[int, float]]:
    codes = ",".join(code7_values)
    population: dict[str, dict[int, float]] = {code: {} for code in code7_values}

    estimates_url = (
        "https://apisidra.ibge.gov.br/values/t/6579/n6/"
        f"{codes}/v/9324/p/2020,2021,2024,2025?formato=json"
    )
    for row in sidra_json(estimates_url)[1:]:
        population[str(row["D1C"])][int(row["D3C"])] = float(row["V"])

    census_url = f"https://apisidra.ibge.gov.br/values/t/4714/n6/{codes}/v/93/p/2022?formato=json"
    for row in sidra_json(census_url)[1:]:
        population[str(row["D1C"])][2022] = float(row["V"])

    for code in code7_values:
        pop_2022 = population[code][2022]
        pop_2024 = population[code][2024]
        population[code][2023] = pop_2022 + ((pop_2024 - pop_2022) / 2)

    return population


def clone_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.number_format:
        target.number_format = source.number_format


def recreate_sheet(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    template = wb["Capitais - AIH"]
    ws = wb.create_sheet(sheet_name)

    for row in range(1, 34):
        ws.row_dimensions[row].height = template.row_dimensions[row].height
        for col in range(1, 9):
            clone_cell_style(template.cell(row=row, column=col), ws.cell(row=row, column=col))

    for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = template.column_dimensions[col_letter].width

    ws.freeze_panes = template.freeze_panes
    ws.sheet_view.showGridLines = template.sheet_view.showGridLines
    ws.merge_cells("A1:G1")
    ws.merge_cells("B2:G2")
    return ws


def populate_sheet(ws, config: dict[str, str], rows: list[dict[str, object]], code_map, population) -> None:
    ws["A1"] = config["title"]
    ws["A2"] = "Capital"
    ws["B2"] = config["label"]
    for index, year in enumerate(YEARS, start=2):
        ws.cell(row=3, column=index, value=year)

    for row_index, item in enumerate(rows, start=4):
        ws.cell(row=row_index, column=1, value=item["location"])
        code7 = code_map[item["code6"]]
        for col_index, year in enumerate(YEARS, start=2):
            raw_value = item["raw"][year]
            pop_value = population[code7][year]
            ws.cell(row=row_index, column=col_index, value=round(raw_value / pop_value * 1000, 1))

    notes_start = 34
    notes = [
        "Fonte e notas",
        f"Fonte da produção no PDF: {config['source']}.",
        "Fonte dos dados: Sistema de Informações Hospitalares do SUS - SIH/SUS / DATASUS.",
        "Período: janeiro a dezembro de 2020 a 2025. Dados extraídos em 04/07/2026; situação da base em 11/06/2026 às 10:29, conforme PDF.",
        "População: IBGE/SIDRA tabela 6579, variável 9324, para 2020, 2021, 2024 e 2025; Censo Demográfico 2022/SIDRA tabela 4714, variável 93, para 2022.",
        "Para 2023, foi usada interpolação linear entre a população do Censo 2022 e a estimativa 2024, mantendo o critério das abas de capitais anteriores.",
        "Razão calculada como (quantidade aprovada / população) x 1.000, para comparação entre capitais.",
    ]
    for offset, note in enumerate(notes):
        ws.cell(row=notes_start + offset, column=1, value=note)


def main() -> None:
    all_rows = {config["sheet"]: parse_pdf_rows(config["pdf"]) for config in INPUTS}
    code6_values = {row["code6"] for rows in all_rows.values() for row in rows}
    code_map = fetch_code_map(code6_values)
    population = fetch_population([code_map[code6] for code6 in sorted(code6_values)])

    wb = load_workbook(WORKBOOK_PATH)
    for config in INPUTS:
        ws = recreate_sheet(wb, config["sheet"])
        populate_sheet(ws, config, all_rows[config["sheet"]], code_map, population)

    wb.save(WORKBOOK_PATH)

    summary = {
        "updatedAt": datetime.now().isoformat(timespec="seconds"),
        "workbook": str(WORKBOOK_PATH),
        "sheets": [config["sheet"] for config in INPUTS],
        "rowsPerSheet": {sheet: len(rows) for sheet, rows in all_rows.items()},
        "rioSamples": {},
    }
    for config in INPUTS:
        rows = all_rows[config["sheet"]]
        rio = next(row for row in rows if row["location"] == "Rio de Janeiro (RJ)")
        code7 = code_map[rio["code6"]]
        summary["rioSamples"][config["sheet"]] = {
            str(year): round(rio["raw"][year] / population[code7][year] * 1000, 1)
            for year in YEARS
        }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
