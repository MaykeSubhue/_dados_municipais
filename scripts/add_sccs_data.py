from __future__ import annotations

import json
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = ROOT / "tmp" / "demanda_comparacao_municipios.xlsx"
DOWNLOADS = Path.home() / "Downloads"
SOURCE_GLOB = "CONSOLIDADO SUPER CENTRO BANCO DE DADOS*.xlsx"

YEAR = 2025


TARGETS = {
    "production": {
        "source_sheet": "Por 100 mil habitantes",
        "target_sheet": "Produção por 100mil habitantes",
        "title": "Produção total por 100 mil habitantes - Capitais do Brasil e SCCS, 2025",
        "location_header": "Órgão / unidade",
        "label": "Procedimentos por 100 mil hab.",
    },
    "oci": {
        "source_sheet": "OCI por 100 mil habitantes",
        "target_sheet": "OCI por 100 mil habitantes",
        "title": "Oferta de Cuidados Integrados por 100 mil habitantes - Municípios do ERJ e SCCS, 2025",
        "location_header": "Município / unidade",
        "label": "OCI por 100 mil hab.",
    },
}


def find_source_file() -> Path:
    matches = [path for path in DOWNLOADS.glob(SOURCE_GLOB) if not path.name.startswith("~$")]
    if not matches:
        raise FileNotFoundError(f"Arquivo não encontrado em {DOWNLOADS} com padrão {SOURCE_GLOB}")
    return max(matches, key=lambda path: path.stat().st_mtime)


def clean_name(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    fixes = {
        "Mage": "Magé",
    }
    return fixes.get(text, text)


def as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def parse_production(ws) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        position = row[0]
        location = clean_name(row[1])
        value = as_float(row[2])
        if not location or value is None:
            continue
        rows.append({"position": int(position), "location": location, "value": value})
    return rows


def merge_oci_rows(rows: list[dict[str, Any]], item: dict[str, Any]) -> None:
    existing = next((row for row in rows if row["location"] == item["location"]), None)
    if existing is None:
        rows.append(item)
        return
    # The source has duplicated locations in the two ranking blocks. Keep the
    # value with more precision when both blocks refer to the same unit.
    if abs(existing["value"] - item["value"]) < 0.05:
        if len(str(item["value"])) > len(str(existing["value"])):
            existing["value"] = item["value"]


def parse_oci(ws) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    blocks = [(1, 2, 5), (7, 8, 11)]
    for position_col, location_col, value_col in blocks:
        for source_row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            position = source_row[position_col - 1]
            location = clean_name(source_row[location_col - 1])
            value = as_float(source_row[value_col - 1])
            if not location or value is None:
                continue
            merge_oci_rows(rows, {"position": int(position), "location": location, "value": value})
    rows.sort(key=lambda row: row["value"], reverse=True)
    for index, row in enumerate(rows, start=1):
        row["position"] = index
    return rows


def clone_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.number_format:
        target.number_format = source.number_format


def recreate_sheet(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    template = wb["Capitais - AIH"]
    ws = wb.create_sheet(sheet_name)

    for row in range(1, 46):
        ws.row_dimensions[row].height = template.row_dimensions[row].height
        for col in range(1, 10):
            clone_cell_style(template.cell(row=row, column=col), ws.cell(row=row, column=col))

    for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
        ws.column_dimensions[col_letter].width = template.column_dimensions[col_letter].width

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.freeze_panes = template.freeze_panes
    ws.sheet_view.showGridLines = template.sheet_view.showGridLines
    ws.merge_cells("A1:B1")
    return ws


def write_standard_sheet(ws, config: dict[str, str], rows: list[dict[str, Any]]) -> None:
    ws["A1"] = config["title"]
    ws["A2"] = config["location_header"]
    ws["B2"] = config["label"]
    ws["B3"] = YEAR

    for row_index, item in enumerate(rows, start=4):
        ws.cell(row=row_index, column=1, value=item["location"])
        ws.cell(row=row_index, column=2, value=item["value"])
        ws.cell(row=row_index, column=2).number_format = "#,##0.00"

    notes = [
        "Fonte e notas",
        "Fonte: CONSOLIDADO SUPER CENTRO BANCO DE DADOS E ANÁLISES.xlsx.",
        "Período: janeiro a dezembro de 2025, conforme metodologia do arquivo de origem.",
        "Produção: total da base tratada dividido pela população de referência, multiplicado por 100 mil.",
        "População: base populacional estimada para 2025, segundo IBGE, conforme metodologia do arquivo de origem.",
        "SCCS: Super Centro Carioca de Saúde; na produção total, Rio de Janeiro aparece sem SCCS quando indicado na fonte.",
    ]
    for offset, note in enumerate(notes, start=6):
        ws.cell(row=offset, column=9, value=note)


def main() -> None:
    source_path = find_source_file()
    source_wb = load_workbook(source_path, data_only=True, read_only=True)
    datasets = {
        TARGETS["production"]["target_sheet"]: parse_production(source_wb[TARGETS["production"]["source_sheet"]]),
        TARGETS["oci"]["target_sheet"]: parse_oci(source_wb[TARGETS["oci"]["source_sheet"]]),
    }

    wb = load_workbook(WORKBOOK_PATH)
    for key, config in TARGETS.items():
        target_sheet = config["target_sheet"]
        ws = recreate_sheet(wb, target_sheet)
        write_standard_sheet(ws, config, datasets[target_sheet])
    wb.save(WORKBOOK_PATH)

    summary = {
        "updatedAt": datetime.now().isoformat(timespec="seconds"),
        "source": str(source_path),
        "workbook": str(WORKBOOK_PATH),
        "sheets": {sheet: len(rows) for sheet, rows in datasets.items()},
        "sccsValues": {
            sheet: next((row["value"] for row in rows if row["location"] == "SCCS"), None)
            for sheet, rows in datasets.items()
        },
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
